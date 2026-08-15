#!/usr/bin/env python3
"""Build collection/shot_list.xlsx from the generated CSVs.

Five sheets. The Balance Check sheet uses LIVE formulas, not baked-in numbers, so the
counts recompute as you edit the Shot List -- if you swap a gesture by hand, the totals
turn red immediately instead of drifting silently.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
COLLECTION = HERE.parent / "collection"

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
GROUP_FILLS = {
    "fold0": PatternFill("solid", fgColor="E8F0FE"),
    "fold1": PatternFill("solid", fgColor="E6F4EA"),
    "fold2": PatternFill("solid", fgColor="FEF7E0"),
    "test": PatternFill("solid", fgColor="FCE8E6"),
}
OK_FILL = PatternFill("solid", fgColor="CEEAD6")
BAD_FILL = PatternFill("solid", fgColor="F5C6C4")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

CLASSES = ["thumbout", "openhand", "closedhand"]
GROUPS = ["fold0", "fold1", "fold2", "test"]


def read(name):
    with (COLLECTION / name).open() as fh:
        return list(csv.DictReader(fh))


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_shots(wb, shots):
    ws = wb.create_sheet("Shot List")
    cols = ["photo_id", "env_id", "slot", "seq_in_env", "filename", "group",
            "triplet_type", "left_class", "right_class", "hand_rotation_deg",
            "expected_label_sequence", "annotations_expected",
            "captured", "mask_done", "qa_pass", "notes"]
    headers = ["Photo", "Env", "Slot", "#", "Filename", "Group", "Triplet",
               "LEFT hand", "RIGHT hand", "Rotation", "Expected labels", "N",
               "Shot?", "Mask?", "QA?", "Notes"]
    ws.append(headers)
    for s in shots:
        ws.append([int(s[c]) if c in ("seq_in_env", "hand_rotation_deg", "annotations_expected")
                   else s[c] for c in cols])

    style_header(ws, len(headers))
    for r, s in enumerate(shots, start=2):
        fill = GROUP_FILLS[s["group"]]
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            cell.fill = fill
            if c in (8, 9):
                cell.font = Font(bold=True, size=11)
            if c in (13, 14, 15):
                cell.alignment = Alignment(horizontal="center")
    autosize(ws, [8, 6, 6, 4, 16, 8, 9, 13, 13, 9, 15, 4, 7, 7, 7, 30])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}61"
    return ws


def sheet_balance(wb):
    """Live COUNTIFS against 'Shot List'. Column F=group, H=left, I=right."""
    ws = wb.create_sheet("Balance Check")
    S = "'Shot List'"

    ws["A1"] = "LIVE CHECK — every green cell must stay green. Edit the Shot List and these recompute."
    ws["A1"].font = Font(bold=True, size=12)

    # --- total annotations per class, per group ---
    ws["A3"] = "Annotations per class (target 10 per group, 40 total)"
    ws["A3"].font = Font(bold=True)
    ws.append([])
    hdr = 4
    ws.cell(row=hdr, column=1, value="Group")
    for j, cls in enumerate(CLASSES, start=2):
        ws.cell(row=hdr, column=j, value=cls)
    ws.cell(row=hdr, column=5, value="Photos")
    ws.cell(row=hdr, column=6, value="Annotations")

    for i, g in enumerate(GROUPS):
        r = hdr + 1 + i
        ws.cell(row=r, column=1, value=g)
        for j, cls in enumerate(CLASSES, start=2):
            col = get_column_letter(j)
            ws.cell(row=r, column=j, value=(
                f'=COUNTIFS({S}!$F$2:$F$61,$A{r},{S}!$H$2:$H$61,{col}${hdr})'
                f'+COUNTIFS({S}!$F$2:$F$61,$A{r},{S}!$I$2:$I$61,{col}${hdr})'
            ))
        ws.cell(row=r, column=5, value=f'=COUNTIF({S}!$F$2:$F$61,$A{r})')
        ws.cell(row=r, column=6, value=f'=SUM(B{r}:D{r})')

    tot = hdr + 5
    ws.cell(row=tot, column=1, value="TOTAL").font = Font(bold=True)
    for j in range(2, 7):
        c = get_column_letter(j)
        ws.cell(row=tot, column=j, value=f'=SUM({c}{hdr+1}:{c}{hdr+4})').font = Font(bold=True)

    ws.conditional_formatting.add(f"B{hdr+1}:D{hdr+4}",
        CellIsRule(operator="equal", formula=["10"], fill=OK_FILL))
    ws.conditional_formatting.add(f"B{hdr+1}:D{hdr+4}",
        CellIsRule(operator="notEqual", formula=["10"], fill=BAD_FILL))
    ws.conditional_formatting.add(f"E{hdr+1}:E{hdr+4}",
        CellIsRule(operator="equal", formula=["15"], fill=OK_FILL))
    ws.conditional_formatting.add(f"E{hdr+1}:E{hdr+4}",
        CellIsRule(operator="notEqual", formula=["15"], fill=BAD_FILL))
    ws.conditional_formatting.add(f"B{tot}:D{tot}",
        CellIsRule(operator="equal", formula=["40"], fill=OK_FILL))
    ws.conditional_formatting.add(f"B{tot}:D{tot}",
        CellIsRule(operator="notEqual", formula=["40"], fill=BAD_FILL))

    # --- per side ---
    base = tot + 3
    ws.cell(row=base, column=1, value="Per hand side (target 5 per class per side, per group)").font = Font(bold=True)
    h = base + 1
    ws.cell(row=h, column=1, value="Group")
    for j, cls in enumerate(CLASSES, start=2):
        ws.cell(row=h, column=j, value=f"{cls} L")
    for j, cls in enumerate(CLASSES, start=5):
        ws.cell(row=h, column=j, value=f"{cls} R")

    for i, g in enumerate(GROUPS):
        r = h + 1 + i
        ws.cell(row=r, column=1, value=g)
        for j, cls in enumerate(CLASSES, start=2):
            ws.cell(row=r, column=j, value=(
                f'=COUNTIFS({S}!$F$2:$F$61,$A{r},{S}!$H$2:$H$61,"{cls}")'))
        for j, cls in enumerate(CLASSES, start=5):
            ws.cell(row=r, column=j, value=(
                f'=COUNTIFS({S}!$F$2:$F$61,$A{r},{S}!$I$2:$I$61,"{cls}")'))
    ws.conditional_formatting.add(f"B{h+1}:G{h+4}",
        CellIsRule(operator="equal", formula=["5"], fill=OK_FILL))
    ws.conditional_formatting.add(f"B{h+1}:G{h+4}",
        CellIsRule(operator="notEqual", formula=["5"], fill=BAD_FILL))

    # --- rotation ---
    base2 = h + 6
    ws.cell(row=base2, column=1, value="Rotation balance (target 20 photos at each)").font = Font(bold=True)
    h2 = base2 + 1
    ws.cell(row=h2, column=1, value="Rotation")
    ws.cell(row=h2, column=2, value="Photos")
    for i, rot in enumerate([0, 30, 60]):
        r = h2 + 1 + i
        ws.cell(row=r, column=1, value=rot)
        ws.cell(row=r, column=2, value=f'=COUNTIF({S}!$J$2:$J$61,$A{r})')
    ws.conditional_formatting.add(f"B{h2+1}:B{h2+3}",
        CellIsRule(operator="equal", formula=["20"], fill=OK_FILL))
    ws.conditional_formatting.add(f"B{h2+1}:B{h2+3}",
        CellIsRule(operator="notEqual", formula=["20"], fill=BAD_FILL))

    # --- progress ---
    base3 = h2 + 5
    ws.cell(row=base3, column=1, value="Progress").font = Font(bold=True)
    for i, (label, col) in enumerate([("Photos shot", "M"), ("Masks done", "N"), ("QA passed", "O")]):
        r = base3 + 1 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=f'=COUNTA({S}!${col}$2:${col}$61)')
        ws.cell(row=r, column=3, value="of 60")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = BOX
    autosize(ws, [26, 14, 14, 14, 14, 14, 14])
    return ws


def sheet_envs(wb, envs):
    ws = wb.create_sheet("Environments")
    cols = ["env_id", "group", "slot", "room", "background_desc", "light_class",
            "light_source", "lux_target", "camera_distance_cm", "camera_tilt_deg",
            "sleeve", "accessory", "accessory_side", "difficulty", "n_photos",
            "photo_ids", "session_block", "time_of_day", "room_confirmed", "notes"]
    headers = ["Env", "Group", "Slot", "Room (CONFIRM)", "Background", "Light",
               "Light source", "Lux", "Distance cm", "Tilt deg", "Sleeve",
               "Accessory", "Side", "Difficulty", "Photos", "Photo IDs",
               "Session", "Time", "Confirmed?", "Notes"]
    ws.append(headers)
    for e in envs:
        ws.append([int(e[c]) if c in ("lux_target", "camera_distance_cm",
                                      "camera_tilt_deg", "n_photos") else e[c] for c in cols])
    style_header(ws, len(headers))
    for r, e in enumerate(envs, start=2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            cell.fill = GROUP_FILLS[e["group"]]
            cell.alignment = WRAP
    autosize(ws, [6, 8, 6, 22, 28, 10, 34, 7, 11, 9, 26, 11, 12, 10, 7, 10, 10, 9, 11, 34])
    return ws


def sheet_cards(wb, envs, shots):
    """One printable block per environment: set up, then shoot straight down the list."""
    ws = wb.create_sheet("Setup Cards")
    r = 1
    by_env = {}
    for s in shots:
        by_env.setdefault(s["env_id"], []).append(s)

    for e in envs:
        eid = e["env_id"]
        ws.cell(row=r, column=1, value=f'{eid}  —  {e["room"]}  ({e["group"]}, slot {e["slot"]})')
        ws.cell(row=r, column=1).font = Font(bold=True, size=13, color="FFFFFF")
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = HEAD_FILL
        r += 1

        for label, val in [
            ("Background", e["background_desc"]),
            ("Lighting", f'{e["light_source"]}  (~{e["lux_target"]} lux, {e["light_class"]})'),
            ("Camera", f'{e["camera_distance_cm"]} cm from hands, screen tilt {e["camera_tilt_deg"]}°'),
            ("Wear", f'{e["sleeve"]}' + (f'  +  {e["accessory"]} on the {e["accessory_side"]} hand'
                                         if e["accessory"] != "none" else "")),
            ("Session", f'{e["session_block"]}, {e["time_of_day"]}'),
        ]:
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val)
            r += 1

        r += 1
        for j, h in enumerate(["Photo", "File", "LEFT hand", "RIGHT hand", "Rotation", "Shot?"], start=1):
            cell = ws.cell(row=r, column=j, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor="D9D9D9")
            cell.border = BOX
        r += 1
        for s in by_env[eid]:
            vals = [s["photo_id"], s["filename"], s["left_class"], s["right_class"],
                    f'{s["hand_rotation_deg"]}°', ""]
            for j, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.border = BOX
                if j in (3, 4):
                    cell.font = Font(bold=True, size=11)
            r += 1
        r += 2

    autosize(ws, [14, 20, 16, 16, 11, 9, 20])
    return ws


def sheet_protocol(wb):
    ws = wb.create_sheet("Protocol")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 108

    blocks = [
        ("THE RULES THAT MATTER MOST", [
            "Lock the camera's exposure, white balance and focus BEFORE the first photo of an "
            "environment, and do not let them change between the photos of that environment. If "
            "the camera re-meters between an open hand and a fist, brightness becomes a clue to "
            "the gesture and the whole environment is wasted. This cannot be fixed afterwards.",
            "Keep the 18% grey card taped in the same corner of the frame for all 60 photos. It is "
            "how we prove the exposure stayed locked.",
            "Do not touch the laptop between the photos of one environment. Use the delayed "
            "trigger so both hands are free.",
            "Both hands must be fully inside the frame, not overlapping each other, with a gap of "
            "about one hand-width between them.",
        ]),
        ("SETTING UP AN ENVIRONMENT", [
            "Position the laptop at the distance and screen tilt on the setup card.",
            "Set the lighting exactly as the card says. Nothing else on.",
            "Put on the sleeves and accessory the card specifies.",
            "Tape the grey card into the top-left of the frame.",
            "Lock exposure / white balance / focus.",
            "Shoot the photos in the order on the card. Do not reorder them.",
        ]),
        ("THE GESTURES", [
            "open hand — all five fingers extended and spread, palm to the camera.",
            "closed hand — a fist, fingers curled in, thumb resting against the index finger.",
            "thumb out — fingers curled into a fist, thumb extended clearly away from the fist.",
            "Rotation is in the plane of the screen: the palm stays facing the camera and the whole "
            "hand rotates. 0° = fingers point straight up. 30° and 60° = tilted by that much. "
            "Rotate both hands by the same amount.",
        ]),
        ("WHAT MAKES A PHOTO INVALID — reshoot it", [
            "Either hand is cut off by the frame edge.",
            "The hands overlap or touch each other.",
            "Motion blur — hold still through the shutter delay.",
            "The gesture is ambiguous (a half-closed hand is neither open nor closed).",
            "The grey card is missing, obscured, or blown out.",
            "The exposure visibly changed from the previous photo in the same environment.",
        ]),
        ("ORDER OF WORK", [
            "Confirm every room on the Environments sheet exists and is free. Mark Confirmed?.",
            "Measure the webcam's field of view with a ruler at 55 cm and 75 cm before locking the "
            "distances. If the frame is narrower than expected, add the same amount to all three "
            "distances rather than changing them individually.",
            "Shoot E01 first as a pilot, all the way through masks and QA, before shooting anything "
            "else. It is cheaper to find a problem in 3 photos than in 60.",
            "Never shoot a test environment (E10, E11, E12) in the same session as a training one.",
            "Shoot the dark, hard environments first in each session, while you are fresh.",
        ]),
        ("WHY THE DESIGN LOOKS LIKE THIS", [
            "Every environment contains each gesture exactly the same number of times, on each "
            "hand. So the background, the lighting and the sleeve carry no information about which "
            "gesture is being made, and the model cannot take a shortcut.",
            "All twelve places are split so that a place is only ever in one fold. The three test "
            "places appear nowhere in training, which is what lets us claim the model works in a "
            "room it has never seen.",
            "The four groups match on distance, tilt, lighting, sleeve and difficulty, so when the "
            "three fold scores differ, the difference is caused by the places themselves and "
            "nothing else.",
        ]),
    ]

    r = 1
    for title, items in blocks:
        c = ws.cell(row=r, column=2, value=title)
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = HEAD_FILL
        r += 1
        for it in items:
            ws.cell(row=r, column=1, value="•").alignment = Alignment(horizontal="center", vertical="top")
            cell = ws.cell(row=r, column=2, value=it)
            cell.alignment = WRAP
            ws.row_dimensions[r].height = max(15, 13 * (len(it) // 95 + 1))
            r += 1
        r += 1
    return ws


if __name__ == "__main__":
    envs = read("environments.csv")
    shots = read("shotlist.csv")

    wb = Workbook()
    wb.remove(wb.active)
    sheet_shots(wb, shots)
    sheet_balance(wb)
    sheet_cards(wb, envs, shots)
    sheet_envs(wb, envs)
    sheet_protocol(wb)

    out = COLLECTION / "shot_list.xlsx"
    wb.save(out)
    print(f"wrote {out}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")
